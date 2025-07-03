import tkinter as tk
from tkinter import ttk, filedialog, messagebox, colorchooser
from datetime import datetime, timedelta
import json
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os
from tkcalendar import Calendar # Import Calendar widget
import threading # For running AI generation in a separate thread

# --- Google Generative AI Imports ---
# You need to install these: pip install google-generativeai langchain-google-genai
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_core.messages import HumanMessage

class GanttChartApp:
    def __init__(self, master):
        self.master = master
        master.title("Python Gantt Chart Tool")
        master.geometry("1200x800") # Increased initial window size

        self.tasks = []
        self.edit_index = -1
        self.current_sort_criteria = None
        self.file_dialog_initial_dir = os.getcwd() # Store last used directory

        # For Undo/Redo
        self.history = []
        self.history_index = -1

        # For canvas panning
        self.canvas_x_offset = 0
        self.canvas_y_offset = 0
        self._drag_data = {"x": 0, "y": 0, "item": None}

        # --- Google Generative AI Configuration ---
        # WARNING: Embedding API keys directly in client-side applications is generally NOT secure.
        # For production, consider environment variables or a secure backend service.
        # Ensure you set the GOOGLE_API_KEY environment variable before running this application.
        # Example (Linux/macOS): export GOOGLE_API_KEY="YOUR_ACTUAL_API_KEY"
        # Example (Windows cmd): set GOOGLE_API_KEY=YOUR_ACTUAL_API_KEY
        self.GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY") 

        if not self.GOOGLE_API_KEY:
            messagebox.showwarning("API Key Missing", 
                                 "GOOGLE_API_KEY environment variable is not set.\n"
                                 "AI features will be disabled. Please set your API key.")
            self.llm = None
        else:
            try:
                os.environ["GOOGLE_API_KEY"] = self.GOOGLE_API_KEY
                self.llm = ChatGoogleGenerativeAI(model="gemini-1.5-flash", temperature=0.1, max_output_tokens=7500)
                print("Google Generative AI LLM initialized successfully within Tkinter app.")
            except Exception as e:
                self.llm = None
                messagebox.showerror("AI Initialization Error", 
                                     f"Failed to initialize Google Generative AI: {e}\n"
                                     "AI features will be disabled. Please check your API key and network connection.")
                print(f"Error initializing Google LLM: {e}")

        # --- Apply professional styling ---
        self.style = ttk.Style()
        self.configure_styles()

        # 1. Create all GUI elements first
        self.create_widgets() 
        # Force Tkinter to update and make widgets fully available
        self.master.update_idletasks()
        
        # 2. Load tasks from file (this populates self.tasks)
        self.load_tasks() 

        # 3. Update dynamic elements (like epic filter options) and history.
        self.update_epic_filter_options() 
        self._save_history() # Save initial state after widgets are created and tasks loaded

        # 4. Render chart initially
        self.render_chart(self.filter_epic_var.get().strip()) # Pass initial filter value

        # 5. Bind canvas resizing and panning
        self.chart_canvas.bind("<Configure>", self.on_canvas_resize)
        self.chart_canvas.bind("<ButtonPress-1>", self.on_canvas_press)
        self.chart_canvas.bind("<B1-Motion>", self.on_canvas_drag)
        self.chart_canvas.bind("<ButtonRelease-1>", self.on_canvas_release)

    def configure_styles(self):
        # Define professional color palette
        self.colors = {
            "primary_bg": "#F5F5F5",        # Very light gray
            "secondary_bg": "#E8E8E8",      # Light gray for frames
            "accent_blue": "#003366",       # Deep, rich blue
            "accent_hover": "#004488",      # Slightly lighter blue for hover
            "text_dark": "#333333",         # Dark gray for most text
            "text_light": "#FFFFFF",        # White for text on dark backgrounds
            "border_light": "#CCCCCC",      # Light gray border
            "border_dark": "#999999",       # Darker gray border
            "valid_green": "#28a745",       # Green for valid feedback
            "error_red": "#dc3545",         # Red for error feedback
            "canvas_bg": "#FFFFFF",         # White for canvas background
            "current_week_highlight": "#E6F3FF" # Light blue for current week
        }

        # Define common font
        self.font_family = "Arial"
        self.font_base_size = 9
        self.font_bold = (self.font_family, self.font_base_size, "bold")
        self.font_small = (self.font_family, self.font_base_size - 2)
        self.font_normal = (self.font_family, self.font_base_size)

        # General style for all widgets
        self.style.configure('.', font=self.font_normal, background=self.colors["primary_bg"], foreground=self.colors["text_dark"])
        self.style.map('.', background=[('active', self.colors["primary_bg"])]) # Prevent default active background on some widgets

        # TFrame and TLabelFrame
        self.style.configure('TFrame', background=self.colors["primary_bg"])
        self.style.configure('TLabelFrame', background=self.colors["secondary_bg"], foreground=self.colors["text_dark"],
                             font=(self.font_family, self.font_base_size + 1, "bold"))
        self.style.configure('TLabelframe.Label', background=self.colors["secondary_bg"], foreground=self.colors["text_dark"]) # Ensure label background matches frame

        # TButton
        self.style.configure('TButton', 
                             background=self.colors["accent_blue"], 
                             foreground=self.colors["text_light"],
                             font=self.font_bold,
                             padding=[10, 5]) # Add padding for a softer look
        self.style.map('TButton', 
                       background=[('active', self.colors["accent_hover"]), ('pressed', self.colors["accent_blue"])],
                       foreground=[('active', self.colors["text_light"])])

        # TEntry and TCombobox
        self.style.configure('TEntry', 
                             fieldbackground=self.colors["canvas_bg"], 
                             foreground=self.colors["text_dark"],
                             bordercolor=self.colors["border_light"],
                             lightcolor=self.colors["border_light"],
                             darkcolor=self.colors["border_dark"],
                             padding=[5, 3])
        self.style.configure('TCombobox', 
                             fieldbackground=self.colors["canvas_bg"], 
                             foreground=self.colors["text_dark"],
                             bordercolor=self.colors["border_light"],
                             lightcolor=self.colors["border_light"],
                             darkcolor=self.colors["border_dark"],
                             arrowcolor=self.colors["accent_blue"],
                             padding=[5, 3])
        self.style.map('TCombobox', 
                       fieldbackground=[('readonly', self.colors["canvas_bg"])],
                       selectbackground=[('readonly', self.colors["accent_blue"])],
                       selectforeground=[('readonly', self.colors["text_light"])])
        
        # TLabel
        self.style.configure('TLabel', background=self.colors["primary_bg"], foreground=self.colors["text_dark"])

        # TCheckbutton
        self.style.configure('TCheckbutton', background=self.colors["secondary_bg"], foreground=self.colors["text_dark"])
        self.style.map('TCheckbutton', background=[('active', self.colors["secondary_bg"])]) # Prevent background change on hover

        # Specific styles for AI Assist window buttons
        self.style.configure('AiButton.TButton', 
                             background=self.colors["accent_blue"], 
                             foreground=self.colors["text_light"],
                             font=self.font_bold,
                             padding=[8, 4])
        self.style.map('AiButton.TButton', 
                       background=[('active', self.colors["accent_hover"]), ('pressed', self.colors["accent_blue"])])
        
        self.style.configure('ApplyButton.TButton', 
                             background=self.colors["accent_blue"], 
                             foreground=self.colors["text_light"],
                             font=self.font_bold,
                             padding=[8, 4])
        self.style.map('ApplyButton.TButton', 
                       background=[('active', self.colors["accent_hover"]), ('pressed', self.colors["accent_blue"])])

        # Status bar style
        self.style.configure('StatusBar.TLabel', background=self.colors["secondary_bg"], foreground=self.colors["text_dark"],
                             font=self.font_small, padding=[5, 2])


    def create_widgets(self):
        # --- Task Input Frame ---
        input_frame = ttk.LabelFrame(self.master, text="Task Details", padding="15") # Increased padding
        input_frame.pack(fill=tk.X, padx=15, pady=10) # Increased padx, pady

        # Row 0
        ttk.Label(input_frame, text="Task Name:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.task_name_entry = ttk.Entry(input_frame, width=30)
        self.task_name_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        # AI Assist button for Task Name
        self.ai_assist_button = ttk.Button(input_frame, text="AI Assist", command=self.open_ai_assist_window, style='AiButton.TButton')
        self.ai_assist_button.grid(row=0, column=1, sticky="e", padx=(0, 5))
        if self.llm is None: # Disable AI button if LLM failed to initialize
            self.ai_assist_button.config(state=tk.DISABLED)


        ttk.Label(input_frame, text="Epic #:").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        self.epic_number_entry = ttk.Entry(input_frame, width=15)
        self.epic_number_entry.grid(row=0, column=3, padx=5, pady=5, sticky="ew")

        ttk.Label(input_frame, text="Start Date (YYYY-MM-DD):").grid(row=0, column=4, padx=5, pady=5, sticky="w")
        self.start_date_entry = ttk.Entry(input_frame, width=15)
        self.start_date_entry.grid(row=0, column=5, padx=5, pady=5, sticky="ew")
        self.start_date_entry.config(state='readonly') # Make it read-only
        ttk.Button(input_frame, text="🗓️", command=lambda: self.open_calendar(self.start_date_entry), width=3).grid(row=0, column=5, sticky="e", padx=(0, 5))
        self.start_date_feedback = ttk.Label(input_frame, text="", font=self.font_small, foreground="gray")
        self.start_date_feedback.grid(row=1, column=5, sticky="w", padx=5)
        self.start_date_entry.bind("<FocusOut>", lambda e: self.validate_date_entry(self.start_date_entry, self.start_date_feedback))


        ttk.Label(input_frame, text="End Date (YYYY-MM-DD):").grid(row=0, column=6, padx=5, pady=5, sticky="w")
        self.end_date_entry = ttk.Entry(input_frame, width=15)
        self.end_date_entry.grid(row=0, column=7, padx=5, pady=5, sticky="ew")
        self.end_date_entry.config(state='readonly') # Make it read-only
        ttk.Button(input_frame, text="🗓️", command=lambda: self.open_calendar(self.end_date_entry), width=3).grid(row=0, column=7, sticky="e", padx=(0, 5))
        self.end_date_feedback = ttk.Label(input_frame, text="", font=self.font_small, foreground="gray")
        self.end_date_feedback.grid(row=1, column=7, sticky="w", padx=5)
        self.end_date_entry.bind("<FocusOut>", lambda e: self.validate_date_entry(self.end_date_entry, self.end_date_feedback))

        ttk.Label(input_frame, text="Color:").grid(row=0, column=8, padx=5, pady=5, sticky="w")
        self.selected_color_var = tk.StringVar(value="#000000") # Variable to hold the selected color
        self.color_display_label = tk.Label(input_frame, textvariable=self.selected_color_var, width=10, relief="sunken", borderwidth=1, bg="#000000", fg="white")
        self.color_display_label.grid(row=0, column=9, padx=5, pady=5, sticky="ew")
        self.color_picker_button = ttk.Button(input_frame, text="🎨", command=self.choose_color, width=3)
        self.color_picker_button.grid(row=0, column=9, sticky="e", padx=(0, 5))

        self.add_update_button = ttk.Button(input_frame, text="Add Task", command=self.add_or_update_task)
        self.add_update_button.grid(row=0, column=10, rowspan=2, padx=5, pady=5, sticky="nsew") # Span two rows

        # Row 1 (for Priority, Status, Milestone)
        ttk.Label(input_frame, text="Priority:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.priority_var = tk.StringVar(value="Medium")
        self.priority_combobox = ttk.Combobox(input_frame, textvariable=self.priority_var, values=["Low", "Medium", "High", "Critical"], width=10, state="readonly")
        self.priority_combobox.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(input_frame, text="Status:").grid(row=1, column=2, padx=5, pady=5, sticky="w")
        self.status_var = tk.StringVar(value="Not Started")
        self.status_combobox = ttk.Combobox(input_frame, textvariable=self.status_var, values=["Not Started", "In Progress", "Completed", "Blocked"], width=10, state="readonly")
        self.status_combobox.grid(row=1, column=3, padx=5, pady=5, sticky="ew")

        self.is_milestone_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(input_frame, text="Is Milestone", variable=self.is_milestone_var).grid(row=1, column=8, padx=5, pady=5, sticky="w")


        input_frame.grid_columnconfigure(1, weight=1)
        input_frame.grid_columnconfigure(3, weight=1)
        input_frame.grid_columnconfigure(5, weight=1)
        input_frame.grid_columnconfigure(7, weight=1)
        input_frame.grid_columnconfigure(9, weight=1)

        # --- Filter Frame ---
        filter_frame = ttk.LabelFrame(self.master, text="Filter", padding="15") # Increased padding
        filter_frame.pack(fill=tk.X, padx=15, pady=10) # Increased padx, pady

        ttk.Label(filter_frame, text="Filter by Epic #:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.filter_epic_var = tk.StringVar()
        self.filter_epic_combobox = ttk.Combobox(filter_frame, textvariable=self.filter_epic_var, width=30)
        self.filter_epic_combobox.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        self.filter_epic_combobox.bind("<<ComboboxSelected>>", self.apply_filter)
        self.filter_epic_combobox.bind("<KeyRelease>", self.apply_filter) # Allow typing to filter options
        filter_frame.grid_columnconfigure(1, weight=1)

        ttk.Label(filter_frame, text="Filter Start Date:").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        self.filter_start_date_entry = ttk.Entry(filter_frame, width=15)
        self.filter_start_date_entry.grid(row=0, column=3, padx=5, pady=5, sticky="ew")
        self.filter_start_date_entry.config(state='readonly')
        ttk.Button(filter_frame, text="🗓️", command=lambda: self.open_calendar(self.filter_start_date_entry, self.apply_filter), width=3).grid(row=0, column=3, sticky="e", padx=(0,5))

        ttk.Label(filter_frame, text="Filter End Date:").grid(row=0, column=4, padx=5, pady=5, sticky="w")
        self.filter_end_date_entry = ttk.Entry(filter_frame, width=15)
        self.filter_end_date_entry.grid(row=0, column=5, padx=5, pady=5, sticky="ew")
        self.filter_end_date_entry.config(state='readonly')
        ttk.Button(filter_frame, text="🗓️", command=lambda: self.open_calendar(self.filter_end_date_entry, self.apply_filter), width=3).grid(row=0, column=5, sticky="e", padx=(0,5))
        
        ttk.Button(filter_frame, text="Clear Filters", command=self.clear_filters).grid(row=0, column=6, padx=5, pady=5)

        filter_frame.grid_columnconfigure(3, weight=1)
        filter_frame.grid_columnconfigure(5, weight=1)

        # --- Sort Frame ---
        sort_frame = ttk.LabelFrame(self.master, text="Sort", padding="15") # Increased padding
        sort_frame.pack(fill=tk.X, padx=15, pady=10) # Increased padx, pady

        ttk.Label(sort_frame, text="Sort by:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        ttk.Button(sort_frame, text="Task Name", command=lambda: self.sort_tasks('name')).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(sort_frame, text="Start Date", command=lambda: self.sort_tasks('start_date')).grid(row=0, column=2, padx=5, pady=5)
        ttk.Button(sort_frame, text="End Date", command=lambda: self.sort_tasks('end_date')).grid(row=0, column=3, padx=5, pady=5)
        ttk.Button(sort_frame, text="Priority", command=lambda: self.sort_tasks('priority')).grid(row=0, column=4, padx=5, pady=5)
        ttk.Button(sort_frame, text="Status", command=lambda: self.sort_tasks('status')).grid(row=0, column=5, padx=5, pady=5)


        # --- Gantt Chart Canvas with Scrollbars ---
        canvas_frame = ttk.Frame(self.master)
        canvas_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=10) # Increased padx, pady

        self.chart_canvas = tk.Canvas(canvas_frame, bg=self.colors["canvas_bg"], borderwidth=2, relief="groove", highlightbackground=self.colors["border_light"])
        self.chart_canvas.grid(row=0, column=0, sticky="nsew")

        self.v_scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=self.chart_canvas.yview)
        self.v_scrollbar.grid(row=0, column=1, sticky="ns")
        self.chart_canvas.config(yscrollcommand=self.v_scrollbar.set)

        self.h_scrollbar = ttk.Scrollbar(canvas_frame, orient="horizontal", command=self.chart_canvas.xview)
        self.h_scrollbar.grid(row=1, column=0, sticky="ew")
        self.chart_canvas.config(xscrollcommand=self.h_scrollbar.set)

        canvas_frame.grid_rowconfigure(0, weight=1)
        canvas_frame.grid_columnconfigure(0, weight=1)

        # --- Export/Import/Undo/Redo Buttons ---
        button_frame = ttk.Frame(self.master, padding="15") # Increased padding
        button_frame.pack(fill=tk.X, padx=15, pady=10) # Increased padx, pady

        ttk.Button(button_frame, text="Save Data (JSON)", command=self.save_tasks_to_file).grid(row=0, column=0, padx=5, pady=5)
        ttk.Button(button_frame, text="Load Data (JSON)", command=self.load_tasks_from_file).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(button_frame, text="Export to Excel", command=self.export_to_excel).grid(row=0, column=2, padx=5, pady=5)
        ttk.Button(button_frame, text="Export to PNG", command=self.export_to_png).grid(row=0, column=3, padx=5, pady=5)

        self.undo_button = ttk.Button(button_frame, text="Undo", command=self.undo_action, state=tk.DISABLED)
        self.undo_button.grid(row=0, column=4, padx=5, pady=5)
        self.redo_button = ttk.Button(button_frame, text="Redo", command=self.redo_action, state=tk.DISABLED)
        self.redo_button.grid(row=0, column=5, padx=5, pady=5)
        
        # Theme selection
        ttk.Label(button_frame, text="Theme:").grid(row=0, column=6, padx=5, pady=5, sticky="e")
        self.theme_var = tk.StringVar()
        self.theme_combobox = ttk.Combobox(button_frame, textvariable=self.theme_var, values=self.master.tk.call("ttk::themes"), state="readonly", width=10)
        self.theme_combobox.grid(row=0, column=7, padx=5, pady=5, sticky="ew")
        self.theme_combobox.set(self.master.tk.call("ttk::style", "theme", "use")) # Set current theme
        self.theme_combobox.bind("<<ComboboxSelected>>", self.change_theme)


        # Configure button frame columns to expand
        for i in range(8):
            button_frame.grid_columnconfigure(i, weight=1)

        # --- Status Bar ---
        self.status_bar = ttk.Label(self.master, text="Ready", relief=tk.SUNKEN, anchor=tk.W, style='StatusBar.TLabel')
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def change_theme(self, event=None):
        self.master.tk.call("ttk::setTheme", self.theme_var.get())
        self.update_status(f"Theme changed to {self.theme_var.get()}")

    def update_status(self, message):
        self.status_bar.config(text=message)

    def on_canvas_resize(self, event):
        # Recalculate and re-render the chart when canvas size changes
        self.render_chart(self.filter_epic_var.get().strip()) # Pass current epic filter value

    def on_canvas_press(self, event):
        self._drag_data["x"] = event.x
        self._drag_data["y"] = event.y
        self._drag_data["item"] = "canvas" # Indicate we are dragging the canvas view

    def on_canvas_drag(self, event):
        if self._drag_data["item"] == "canvas":
            dx = event.x - self._drag_data["x"]
            dy = event.y - self._drag_data["y"]
            
            # Update canvas offsets
            self.canvas_x_offset += dx
            self.canvas_y_offset += dy
            
            # Scroll the canvas view
            self.chart_canvas.xview_scroll(-dx, "pixels")
            self.chart_canvas.yview_scroll(-dy, "pixels")
            
            self._drag_data["x"] = event.x
            self._drag_data["y"] = event.y

    def on_canvas_release(self, event):
        self._drag_data["item"] = None
        # After release, re-render to ensure elements are drawn correctly based on new offsets
        # self.render_chart(self.filter_epic_var.get().strip()) # Not strictly necessary if xview_scroll handles it

    def validate_date_entry(self, entry_field, feedback_label):
        date_str = entry_field.get().strip()
        if not date_str:
            feedback_label.config(text="")
            return True
        try:
            datetime.strptime(date_str, "%Y-%m-%d")
            feedback_label.config(text="Valid", foreground=self.colors["valid_green"])
            return True
        except ValueError:
            feedback_label.config(text="Invalid Date!", foreground=self.colors["error_red"])
            return False

    def open_calendar(self, entry_field, callback=None):
        def set_date():
            selected_date = cal.selection_get()
            entry_field.config(state='normal')
            entry_field.delete(0, tk.END)
            entry_field.insert(0, selected_date.strftime("%Y-%m-%d"))
            entry_field.config(state='readonly')
            top.destroy()
            if callback:
                callback() # Call the provided callback (e.g., apply_filter)

        top = tk.Toplevel(self.master)
        top.title("Select Date")
        top.transient(self.master) # Make it appear on top of the main window
        top.grab_set() # Make it modal

        # Position the calendar near the entry field
        # Update id to current
        entry_field.update_idletasks() # Ensure widget position is up-to-date
        x = self.master.winfo_x() + entry_field.winfo_x()
        y = self.master.winfo_y() + entry_field.winfo_y() + entry_field.winfo_height()
        top.geometry(f"+{x}+{y}")

        cal = Calendar(top, selectmode='day',
                       date_pattern='yyyy-mm-dd')
        cal.pack(padx=10, pady=10)

        ttk.Button(top, text="Select", command=set_date, style='TButton').pack(pady=5) # Use TButton style
        self.master.wait_window(top) # Wait until the calendar window is closed

    def choose_color(self):
        color_code = colorchooser.askcolor(title="Choose Task Color", initialcolor=self.selected_color_var.get())
        if color_code[1]: # color_code[1] is the hex string
            self.selected_color_var.set(color_code[1])
            self.color_display_label.config(bg=color_code[1])
            # Adjust text color for contrast
            r, g, b = self.master.winfo_rgb(color_code[1]) # Get RGB tuple (0-65535)
            luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 65535 # Perceived brightness
            self.color_display_label.config(fg="black" if luminance > 0.5 else "white")

    def open_ai_assist_window(self):
        if self.llm is None:
            messagebox.showwarning("AI Not Available", "AI features are disabled due to initialization error.")
            return

        self.ai_assist_window = tk.Toplevel(self.master)
        self.ai_assist_window.title("AI Task Assistant")
        self.ai_assist_window.geometry("600x450")
        self.ai_assist_window.transient(self.master)
        self.ai_assist_window.grab_set()
        self.ai_assist_window.configure(bg=self.colors["primary_bg"]) # Set background for AI window

        # Input Frame
        input_frame = ttk.Frame(self.ai_assist_window, padding="10", style='TFrame')
        input_frame.pack(fill=tk.X)

        ttk.Label(input_frame, text="Task Name:", style='TLabel').grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.ai_task_name_entry = ttk.Entry(input_frame, width=40, style='TEntry')
        self.ai_task_name_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        self.ai_task_name_entry.insert(0, self.task_name_entry.get()) # Pre-fill from main task entry
        input_frame.grid_columnconfigure(1, weight=1)

        ttk.Label(input_frame, text="AI Action:", style='TLabel').grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.ai_action_var = tk.StringVar(value="Expand Description")
        self.ai_action_combobox = ttk.Combobox(input_frame, textvariable=self.ai_action_var, 
                                               values=["Expand Description", "Generate Sub-tasks", "Brainstorm Risks", "Draft Status Update"],
                                               state="readonly", width=30, style='TCombobox')
        self.ai_action_combobox.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

        self.ai_generate_button = ttk.Button(input_frame, text="Generate", command=self.run_ai_generation, style='AiButton.TButton')
        self.ai_generate_button.grid(row=2, column=0, columnspan=2, pady=10)

        self.ai_loading_label = ttk.Label(input_frame, text="", foreground="blue", style='TLabel')
        self.ai_loading_label.grid(row=3, column=0, columnspan=2, pady=5)

        # Output Frame
        output_frame = ttk.Frame(self.ai_assist_window, padding="10", style='TFrame')
        output_frame.pack(fill=tk.BOTH, expand=True)

        self.ai_response_text = tk.Text(output_frame, wrap=tk.WORD, height=10, 
                                        bg=self.colors["canvas_bg"], 
                                        fg=self.colors["text_dark"], 
                                        font=self.font_normal,
                                        relief="sunken", borderwidth=1,
                                        insertbackground=self.colors["accent_blue"]) # Cursor color
        self.ai_response_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5) # Added padx, pady for spacing
        
        # Apply Button
        self.ai_apply_button = ttk.Button(self.ai_assist_window, text="Apply to Main Task", command=self.apply_ai_response, style='ApplyButton.TButton')
        self.ai_apply_button.pack(pady=10)

    def run_ai_generation(self):
        if self.llm is None:
            messagebox.showwarning("AI Not Available", "AI features are disabled due to initialization error.")
            return

        task_name = self.ai_task_name_entry.get().strip()
        ai_action = self.ai_action_var.get()

        if not task_name:
            messagebox.showwarning("Input Error", "Please enter a task name for AI assistance.")
            return

        self.ai_loading_label.config(text="Generating...")
        self.ai_generate_button.config(state=tk.DISABLED)
        self.ai_apply_button.config(state=tk.DISABLED)
        self.ai_response_text.delete(1.0, tk.END)

        # Run AI generation in a separate thread to keep UI responsive
        threading.Thread(target=self._generate_ai_response_threaded, args=(task_name, ai_action)).start()

    def _generate_ai_response_threaded(self, task_name, ai_action):
        generated_text = ""
        try:
            prompt_map = {
                "Expand Description": f"Expand on the task '{task_name}' for a project management context. Provide a detailed description.",
                "Generate Sub-tasks": f"Break down the task '{task_name}' into smaller, actionable sub-tasks. List them as bullet points.",
                "Brainstorm Risks": f"What are potential risks and mitigation strategies for a project task named '{task_name}'? List them clearly.",
                "Draft Status Update": f"Draft a concise status update for a task named '{task_name}'. Assume it's 'In Progress', UI is done, backend is 50%."
            }
            
            prompt = prompt_map.get(ai_action, f"Generate content for the task '{task_name}' related to {ai_action}.")

            messages = [HumanMessage(content=prompt)]
            generated_text = self.llm.invoke(messages).content

        except Exception as e:
            generated_text = f"Error during AI generation: {e}"
            print(f"Error during AI generation: {e}")

        # Update UI on the main thread
        self.master.after(0, self._update_ai_assist_ui, generated_text)

    def _update_ai_assist_ui(self, generated_text):
        self.ai_response_text.delete(1.0, tk.END)
        self.ai_response_text.insert(tk.END, generated_text)
        self.ai_loading_label.config(text="Generation Complete!")
        self.ai_generate_button.config(state=tk.NORMAL)
        self.ai_apply_button.config(state=tk.NORMAL)

    def apply_ai_response(self):
        ai_action = self.ai_action_var.get()
        generated_content = self.ai_response_text.get(1.0, tk.END).strip()

        if not generated_content or generated_content.startswith("Error:"):
            messagebox.showwarning("Apply Error", "No valid content to apply.")
            return

        # Clean up common AI prefixes/suffixes for cleaner insertion
        # This part can be refined based on actual LLM output patterns
        clean_prefixes = [
            f"Detailed description for '{self.ai_task_name_entry.get().strip()}':\n\n",
            f"Sub-tasks for '{self.ai_task_name_entry.get().strip()}':\n\n",
            f"Potential risks for '{self.ai_task_name_entry.get().strip()}':\n\n",
            f"Status Update for '{self.ai_task_name_entry.get().strip()}':\n\n",
            "Here is the detailed description:",
            "Here are the sub-tasks:",
            "Here are the potential risks:",
            "Here is the draft status update:",
            "Certainly, here is the content:",
            "```markdown\n", # Remove markdown code block fences if present
            "```"
        ]
        for prefix in clean_prefixes:
            if generated_content.startswith(prefix):
                generated_content = generated_content[len(prefix):].strip()
            if generated_content.endswith(prefix.strip()): # Also check for trailing fences
                generated_content = generated_content[:-len(prefix.strip())].strip()

        if ai_action == "Expand Description":
            # Replace current task name with the expanded description
            self.task_name_entry.delete(0, tk.END)
            self.task_name_entry.insert(0, generated_content)
            self.update_status("AI-generated description applied.")
        elif ai_action == "Generate Sub-tasks":
            # Append sub-tasks to the current task name/description field
            current_task_name = self.task_name_entry.get()
            self.task_name_entry.delete(0, tk.END)
            # Add a clear separator for readability
            self.task_name_entry.insert(0, f"{current_task_name}\n\n--- AI Suggested Sub-tasks ---\n{generated_content}")
            self.update_status("AI-generated sub-tasks appended to description.")
        elif ai_action == "Brainstorm Risks":
            # Append risks to the current task name/description field
            current_task_name = self.task_name_entry.get()
            self.task_name_entry.delete(0, tk.END)
            self.task_name_entry.insert(0, f"{current_task_name}\n\n--- AI Brainstormed Risks ---\n{generated_content}")
            self.update_status("AI-generated risks appended to description.")
        elif ai_action == "Draft Status Update":
            # Append status update to the current task name/description field
            current_task_name = self.task_name_entry.get()
            self.task_name_entry.delete(0, tk.END)
            self.task_name_entry.insert(0, f"{current_task_name}\n\n--- AI Drafted Status Update ---\n{generated_content}")
            self.update_status("AI-generated status update appended to description.")

        self.ai_assist_window.destroy()


    def add_or_update_task(self):
        name = self.task_name_entry.get().strip()
        epic_number = self.epic_number_entry.get().strip()
        
        # Temporarily enable entry fields to get values
        self.start_date_entry.config(state='normal')
        start_date_str = self.start_date_entry.get().strip()
        self.start_date_entry.config(state='readonly')

        self.end_date_entry.config(state='normal')
        end_date_str = self.end_date_entry.get().strip()
        self.end_date_entry.config(state='readonly')

        color = self.selected_color_var.get() # Get color from the color picker variable
        priority = self.priority_var.get()
        status = self.status_var.get()
        is_milestone = self.is_milestone_var.get()

        # Input Validation
        if not name:
            messagebox.showwarning("Input Error", "Task Name is required.")
            return
        if not self.validate_date_entry(self.start_date_entry, self.start_date_feedback) or \
           not self.validate_date_entry(self.end_date_entry, self.end_date_feedback):
            messagebox.showwarning("Input Error", "Please correct invalid date formats.")
            return
        if not start_date_str or not end_date_str:
            messagebox.showwarning("Input Error", "Start Date and End Date are required.")
            return

        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        except ValueError:
            messagebox.showwarning("Input Error", "Date format must beYYYY-MM-DD.")
            return

        if start_date > end_date:
            messagebox.showwarning("Input Error", "End Date cannot be before Start Date.")
            return

        task = {
            "name": name,
            "epic_number": epic_number,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "color": color,
            "priority": priority,
            "status": status,
            "is_milestone": is_milestone
        }

        if self.edit_index >= 0:
            self.tasks[self.edit_index] = task
            self.edit_index = -1
            self.add_update_button.config(text="Add Task")
        else:
            self.tasks.append(task)

        self._save_history() # Save state for undo/redo

        # Clear input fields (keep epic number as per previous request)
        self.task_name_entry.delete(0, tk.END)
        self.start_date_entry.config(state='normal')
        self.start_date_entry.delete(0, tk.END)
        self.start_date_entry.config(state='readonly')
        self.end_date_entry.config(state='normal')
        self.end_date_entry.delete(0, tk.END)
        self.end_date_entry.config(state='readonly')
        self.selected_color_var.set("#000000") # Reset color to default
        self.color_display_label.config(bg="#000000", fg="white")
        self.priority_var.set("Medium") # Reset priority
        self.status_var.set("Not Started") # Reset status
        self.is_milestone_var.set(False) # Reset milestone checkbox

        # Clear date feedback labels
        self.start_date_feedback.config(text="")
        self.end_date_feedback.config(text="")

        self.save_tasks() # Save to default JSON file
        self.update_epic_filter_options() # Update filter options
        self.render_chart(self.filter_epic_var.get().strip())
        self.update_status(f"Task '{name}' added/updated successfully.")

    def edit_task(self, index):
        self.edit_index = index
        task = self.tasks[index]
        self.task_name_entry.delete(0, tk.END)
        self.task_name_entry.insert(0, task["name"])
        self.epic_number_entry.delete(0, tk.END)
        self.epic_number_entry.insert(0, task["epic_number"])
        
        self.start_date_entry.config(state='normal')
        self.start_date_entry.delete(0, tk.END)
        self.start_date_entry.insert(0, task["start_date"])
        self.start_date_entry.config(state='readonly')

        self.end_date_entry.config(state='normal')
        self.end_date_entry.delete(0, tk.END)
        self.end_date_entry.insert(0, task["end_date"])
        self.end_date_entry.config(state='readonly')

        self.selected_color_var.set(task["color"]) # Set color from task
        self.color_display_label.config(bg=task["color"])
        # Adjust text color for contrast when editing
        r, g, b = self.master.winfo_rgb(task["color"])
        luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 65535
        self.color_display_label.config(fg="black" if luminance > 0.5 else "white")

        self.priority_var.set(task.get("priority", "Medium"))
        self.status_var.set(task.get("status", "Not Started"))
        self.is_milestone_var.set(task.get("is_milestone", False))

        self.add_update_button.config(text="Update Task")
        self.update_status(f"Editing task '{task['name']}'.")

    def delete_task(self, index):
        if messagebox.askyesno("Delete Task", "Are you sure you want to delete this task?"):
            task_name = self.tasks[index]['name']
            del self.tasks[index]
            self._save_history() # Save state for undo/redo
            self.save_tasks()
            self.update_epic_filter_options() # Update filter options
            self.render_chart(self.filter_epic_var.get().strip())
            self.update_status(f"Task '{task_name}' deleted.")

    def clear_filters(self):
        self.filter_epic_var.set("All Epics") # Reset epic filter
        self.filter_start_date_entry.config(state='normal')
        self.filter_start_date_entry.delete(0, tk.END)
        self.filter_start_date_entry.config(state='readonly')
        self.filter_end_date_entry.config(state='normal')
        self.filter_end_date_entry.delete(0, tk.END)
        self.filter_end_date_entry.config(state='readonly')
        self.render_chart(self.filter_epic_var.get().strip()) # Pass the cleared epic filter
        self.update_status("All filters cleared.")

    def apply_filter(self, event=None):
        self.render_chart(self.filter_epic_var.get().strip())
        self.update_status("Filters applied.")

    def sort_tasks(self, criteria):
        self.current_sort_criteria = criteria
        
        # Define priority order for sorting
        priority_order = {"Critical": 4, "High": 3, "Medium": 2, "Low": 1}
        status_order = {"Completed": 4, "In Progress": 3, "Not Started": 2, "Blocked": 1}

        self.tasks.sort(key=lambda t: (
            t["name"] if criteria == 'name' else
            datetime.strptime(t["start_date"], "%Y-%m-%d") if criteria == 'start_date' else
            datetime.strptime(t["end_date"], "%Y-%m-%d") if criteria == 'end_date' else
            priority_order.get(t.get("priority", "Medium"), 0) if criteria == 'priority' else # Use .get for safety
            status_order.get(t.get("status", "Not Started"), 0) # Use .get for safety
        ))
        self.render_chart(self.filter_epic_var.get().strip()) # Re-render with current filter
        self.update_status(f"Tasks sorted by {criteria.replace('_', ' ')}.")

    def render_chart(self, epic_filter_value = ''): # Re-added epic_filter_value parameter
        self.chart_canvas.delete("all")

        # Get current canvas dimensions
        canvas_width = self.chart_canvas.winfo_width()
        canvas_height = self.chart_canvas.winfo_height()

        filtered_tasks = self.tasks
        
        # Apply Epic filter
        if epic_filter_value and epic_filter_value != "All Epics":
            filtered_tasks = [t for t in self.tasks if epic_filter_value.lower() in t.get("epic_number", "").lower()]
        
        # Apply Date Range filter
        filter_start_date_str = self.filter_start_date_entry.get().strip()
        filter_end_date_str = self.filter_end_date_entry.get().strip()

        filter_start_date = None
        filter_end_date = None

        try:
            if filter_start_date_str:
                filter_start_date = datetime.strptime(filter_start_date_str, "%Y-%m-%d").date()
            if filter_end_date_str:
                filter_end_date = datetime.strptime(filter_end_date_str, "%Y-%m-%d").date()
        except ValueError:
            # Invalid filter dates, clear them and re-render without date filter
            self.filter_start_date_entry.config(state='normal'); self.filter_start_date_entry.delete(0, tk.END); self.filter_start_date_entry.config(state='readonly')
            self.filter_end_date_entry.config(state='normal'); self.filter_end_date_entry.delete(0, tk.END); self.filter_end_date_entry.config(state='readonly')
            self.update_status("Invalid filter date format. Date filter cleared.")
            self.master.after(100, self.render_chart, epic_filter_value) # Re-render shortly, passing epic filter
            return

        if filter_start_date or filter_end_date:
            temp_filtered_tasks = []
            for task in filtered_tasks:
                task_start = datetime.strptime(task["start_date"], "%Y-%m-%d").date()
                task_end = datetime.strptime(task["end_date"], "%Y-%m-%d").date()

                if filter_start_date and task_end < filter_start_date:
                    continue # Task ends before filter start
                if filter_end_date and task_start > filter_end_date:
                    continue # Task starts after filter end
                temp_filtered_tasks.append(task)
            filtered_tasks = temp_filtered_tasks


        if not filtered_tasks:
            self.chart_canvas.create_text(canvas_width / 2, canvas_height / 2,
                                          text="No tasks to display with current filters.", fill="gray", font=(self.font_family, 14))
            # Set scroll region to current canvas size if no tasks
            self.chart_canvas.config(scrollregion=(0, 0, canvas_width, canvas_height))
            return

        # Determine overall date range from filtered tasks
        all_start_dates = [datetime.strptime(t["start_date"], "%Y-%m-%d").date() for t in filtered_tasks]
        all_end_dates = [datetime.strptime(t["end_date"], "%Y-%m-%d").date() for t in filtered_tasks]

        min_date = min(all_start_dates)
        max_date = max(all_end_dates)

        # Calculate weeks
        weeks = []
        current_week_start = min_date - timedelta(days=min_date.weekday()) # Start of the week (Monday)
        while current_week_start <= max_date:
            weeks.append(current_week_start)
            current_week_start += timedelta(weeks=1)

        # Chart dimensions
        task_height = 30
        row_gap = 10
        header_height = 60
        left_margin = 180 # Increased for task name + epic + priority/status
        week_width = 80 # Width for each week column
        action_col_width = 80 # Width for action buttons

        # Calculate total chart dimensions
        total_chart_width = left_margin + len(weeks) * week_width + action_col_width
        total_chart_height = header_height + len(filtered_tasks) * (task_height + row_gap) + row_gap

        # Update scroll region
        self.chart_canvas.config(scrollregion=(0, 0, total_chart_width, total_chart_height))

        # Draw header
        self.chart_canvas.create_text(left_margin / 2, header_height / 2, text="Task (Epic)\n[P/S]", font=self.font_bold, anchor="center", fill=self.colors["text_dark"])
        
        today_date = datetime.now().date()
        today_week_start = today_date - timedelta(days=today_date.weekday())

        for i, week_start in enumerate(weeks):
            x_pos = left_margin + i * week_width
            week_label = week_start.strftime("%Y-%m-%d") + "\nW" + str(week_start.isocalendar()[1])
            
            # Highlight current week
            fill_color = self.colors["current_week_highlight"] if week_start == today_week_start else self.colors["canvas_bg"]
            self.chart_canvas.create_rectangle(x_pos, 0, x_pos + week_width, header_height, fill=fill_color, outline=self.colors["border_light"])
            self.chart_canvas.create_text(x_pos + week_width / 2, header_height / 2, text=week_label, font=self.font_small, anchor="center", fill=self.colors["text_dark"])
            self.chart_canvas.create_line(x_pos, 0, x_pos, total_chart_height, fill="lightgray", dash=(2,2)) # Vertical grid line

        self.chart_canvas.create_text(left_margin + len(weeks) * week_width + action_col_width / 2, header_height / 2, text="Actions", font=self.font_bold, anchor="center", fill=self.colors["text_dark"])


        # Draw tasks
        for idx, task in enumerate(filtered_tasks):
            y_pos = header_height + idx * (task_height + row_gap) + row_gap
            task_start_date = datetime.strptime(task["start_date"], "%Y-%m-%d").date()
            task_end_date = datetime.strptime(task["end_date"], "%Y-%m-%d").date()

            # Task Name, Epic, Priority, Status
            task_text = f"{task['name']}"
            if task['epic_number']:
                task_text += f" (E:{task['epic_number']})"
            task_text += f"\n[{task.get('priority', 'M')[0]}/{task.get('status', 'NS')[0]}]" # P/S shorthand
            self.chart_canvas.create_text(left_margin - 5, y_pos + task_height / 2, text=task_text, anchor="e", font=self.font_normal, fill=self.colors["text_dark"])

            # Draw task bar
            for i, week_start in enumerate(weeks):
                week_end = week_start + timedelta(days=6)
                if max(task_start_date, week_start) <= min(task_end_date, week_end):
                    bar_x1 = left_margin + i * week_width
                    bar_x2 = bar_x1 + week_width
                    
                    self.chart_canvas.create_rectangle(bar_x1, y_pos, bar_x2, y_pos + task_height, 
                                                       fill=task["color"], outline=task["color"])
                    
                    # Draw milestone marker if applicable
                    if task.get("is_milestone") and task_start_date >= week_start and task_start_date <= week_end:
                        mid_x = bar_x1 + week_width / 2
                        mid_y = y_pos + task_height / 2
                        # Diamond shape for milestone
                        self.chart_canvas.create_polygon(
                            mid_x, y_pos + 2, # Top
                            bar_x2 - 2, mid_y, # Right
                            mid_x, y_pos + task_height - 2, # Bottom
                            bar_x1 + 2, mid_y, # Left
                            fill="gold", outline="darkgoldenrod", width=1
                        )
                        self.chart_canvas.create_text(mid_x, mid_y, text="★", fill="black", font=(self.font_family, 8, "bold"))


                    # Tooltip (using canvas tag for binding)
                    tooltip_rect_id = self.chart_canvas.create_rectangle(bar_x1, y_pos, bar_x2, y_pos + task_height,
                                                                         fill="", outline="", tags="tooltip_area")
                    self.chart_canvas.tag_bind(tooltip_rect_id, "<Enter>", lambda e, tt=(f"Task: {task['name']}\nEpic: {task['epic_number'] or 'N/A'}\n"
                                    f"Start: {task['start_date']}\nEnd: {task['end_date']}\n"
                                    f"Priority: {task.get('priority', 'N/A')}\nStatus: {task.get('status', 'N/A')}"): self.show_tooltip(e, tt))
                    self.chart_canvas.tag_bind(tooltip_rect_id, "<Leave>", self.hide_tooltip)
                else:
                    self.chart_canvas.create_rectangle(left_margin + i * week_width, y_pos, 
                                                       left_margin + (i + 1) * week_width, y_pos + task_height,
                                                       fill=self.colors["canvas_bg"], outline=self.colors["border_light"])


            # Action buttons (Edit, Delete)
            # Find the original index of the task in the self.tasks list
            original_task_index = self.tasks.index(task)

            edit_button = ttk.Button(self.chart_canvas, text="✏️", command=lambda idx=original_task_index: self.edit_task(idx), width=3)
            delete_button = ttk.Button(self.chart_canvas, text="🗑️", command=lambda idx=original_task_index: self.delete_task(idx), width=3)
            
            # Place buttons on canvas
            self.chart_canvas.create_window(left_margin + len(weeks) * week_width + action_col_width / 2 - 25, y_pos + task_height / 2, window=edit_button)
            self.chart_canvas.create_window(left_margin + len(weeks) * week_width + action_col_width / 2 + 25, y_pos + task_height / 2, window=delete_button)

        # Update scroll region after drawing all elements
        self.chart_canvas.config(scrollregion=self.chart_canvas.bbox(tk.ALL))


    # Tooltip functions
    def show_tooltip(self, event, text):
        x, y = event.x_root, event.y_root # Use root coordinates for tooltip window
        self.tooltip_window = tk.Toplevel(self.master)
        self.tooltip_window.wm_overrideredirect(True) # Remove window decorations
        # Position slightly offset from mouse
        self.tooltip_window.wm_geometry(f"+{x + 15}+{y + 15}")
        label = tk.Label(self.tooltip_window, text=text, background="#FFFFCC", relief="solid", borderwidth=1,
                         font=(self.font_family, 8)) # Use defined font
        label.pack(ipadx=1)

    def hide_tooltip(self, event):
        if hasattr(self, 'tooltip_window'):
            self.tooltip_window.destroy()

    # --- History (Undo/Redo) ---
    def _save_history(self):
        # Clear any "future" history if we're not at the end
        self.history = self.history[:self.history_index + 1]
        self.history.append(json.dumps(self.tasks)) # Store tasks as JSON string
        self.history_index = len(self.history) - 1
        self._update_undo_redo_buttons()
        self.save_tasks() # Also save to file on history change

    def _load_from_history(self):
        try:
            self.tasks = json.loads(self.history[self.history_index])
            self.update_epic_filter_options() # Re-added call
            self.render_chart(self.filter_epic_var.get().strip()) # Pass epic filter value
            self._update_undo_redo_buttons()
            self.save_tasks() # Auto-save the state loaded from history
        except Exception as e:
            messagebox.showerror("History Error", f"Failed to load state from history: {e}")

    def undo_action(self):
        if self.history_index > 0:
            self.history_index -= 1
            self._load_from_history()
            self.update_status("Undo successful.")
        else:
            self.update_status("Nothing to undo.")

    def redo_action(self):
        if self.history_index < len(self.history) - 1:
            self.history_index += 1
            self._load_from_history()
            self.update_status("Redo successful.")
        else:
            self.update_status("Nothing to redo.")

    def _update_undo_redo_buttons(self):
        self.undo_button.config(state=tk.NORMAL if self.history_index > 0 else tk.DISABLED)
        self.redo_button.config(state=tk.NORMAL if self.history_index < len(self.history) - 1 else tk.DISABLED)

    # --- File Operations ---
    def save_tasks(self):
        try:
            with open("gantt_tasks.json", "w") as f:
                json.dump(self.tasks, f, indent=4)
            # No status update here, as _save_history already calls this and updates status
        except Exception as e:
            messagebox.showerror("Auto-Save Error", f"Could not auto-save tasks: {e}")

    def load_tasks(self):
        if os.path.exists("gantt_tasks.json"):
            try:
                with open("gantt_tasks.json", "r") as f:
                    self.tasks = json.load(f)
                self.update_status("Tasks loaded automatically from gantt_tasks.json.")
            except json.JSONDecodeError as e:
                messagebox.showerror("Load Error", f"Could not load tasks (JSON decode error in auto-load): {e}")
                self.tasks = []
            except Exception as e:
                messagebox.showerror("Load Error", f"Could not load tasks (auto-load): {e}")
                self.tasks = []

    def save_tasks_to_file(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Save Gantt Chart Data",
            initialdir=self.file_dialog_initial_dir
        )
        if file_path:
            try:
                with open(file_path, "w") as f:
                    json.dump(self.tasks, f, indent=4)
                messagebox.showinfo("Save Successful", f"Tasks saved to {file_path}")
                self.update_status(f"Tasks saved to {os.path.basename(file_path)}.")
                self.file_dialog_initial_dir = os.path.dirname(file_path) # Update last used directory
            except Exception as e:
                messagebox.showerror("Save Error", f"Failed to save tasks: {e}")

    def load_tasks_from_file(self):
        file_path = filedialog.askopenfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("CSV files", "*.csv"), ("All files", "*.*")],
            title="Load Gantt Chart Data",
            initialdir=self.file_dialog_initial_dir
        )
        if file_path:
            try:
                if file_path.endswith(".json"):
                    with open(file_path, "r") as f:
                        self.tasks = json.load(f)
                    messagebox.showinfo("Load Successful", f"Tasks loaded from {file_path}")
                elif file_path.endswith(".csv"):
                    self.tasks = [] # Clear existing tasks for CSV import
                    with open(file_path, "r", newline='') as f:
                        reader = csv.DictReader(f)
                        for row in reader:
                            task = {
                                "name": row.get("name", "").strip(),
                                "epic_number": row.get("epic_number", "").strip(),
                                "start_date": row.get("start_date", "").strip(),
                                "end_date": row.get("end_date", "").strip(),
                                "color": row.get("color", "#000000").strip(),
                                "priority": row.get("priority", "Medium").strip(),
                                "status": row.get("status", "Not Started").strip(),
                                "is_milestone": row.get("is_milestone", "False").lower() == "true"
                            }
                            try:
                                datetime.strptime(task["start_date"], "%Y-%m-%d")
                                datetime.strptime(task["end_date"], "%Y-%m-%d")
                                self.tasks.append(task)
                            except ValueError:
                                print(f"Skipping invalid task row (date format): {row}")
                                continue
                    messagebox.showinfo("Load Successful", f"Tasks loaded from {file_path}")
                else:
                    messagebox.showwarning("Unsupported File Type", "Please select a JSON (.json) or CSV (.csv) file.")
            except json.JSONDecodeError as e:
                messagebox.showerror("Load Error", f"Failed to load tasks (JSON decode error): {e}")
            except Exception as e:
                messagebox.showerror("Load Error", f"Failed to load tasks: {e}")
            
            # After loading, update history, filter options, and render
            self._save_history() 
            self.update_epic_filter_options() # Re-added call
            self.render_chart(self.filter_epic_var.get().strip())
            self.update_status(f"Tasks loaded from {os.path.basename(file_path)}.")
            self.file_dialog_initial_dir = os.path.dirname(file_path) # Update last used directory

    def update_epic_filter_options(self):
        # Get unique epic numbers from current tasks
        unique_epics = sorted(list(set(t["epic_number"] for t in self.tasks if t["epic_number"])))
        
        # Add "All Epics" option at the beginning
        options = ["All Epics"] + unique_epics
        
        # Update the combobox values
        self.filter_epic_combobox['values'] = options
        
        # Set default selection to "All Epics" if the current selection is not valid
        current_selection = self.filter_epic_var.get()
        if not current_selection or current_selection not in options:
            self.filter_epic_var.set("All Epics")


    def export_to_excel(self):
        if not self.tasks:
            messagebox.showinfo("Export", "No tasks to export to Excel.")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Gantt Chart"

        # Headers
        headers = ["Task", "Epic #", "Start Date", "End Date", "Color", "Priority", "Status", "Is Milestone"]
        
        all_start_dates = [datetime.strptime(t["start_date"], "%Y-%m-%d").date() for t in self.tasks]
        all_end_dates = [datetime.strptime(t["end_date"], "%Y-%m-%d").date() for t in self.tasks]

        min_date = min(all_start_dates)
        max_date = max(all_end_dates)

        weeks = []
        current_week_start = min_date - timedelta(days=min_date.weekday())
        while current_week_start <= max_date:
            weeks.append(current_week_start)
            current_week_start += timedelta(weeks=1)
        
        week_headers = [f"{w.strftime('%Y-%m-%d')} W{w.isocalendar()[1]}" for w in weeks]
        ws.append(headers + week_headers)

        # Task data and coloring
        for r_idx, task in enumerate(self.tasks):
            row_data = [
                task["name"],
                task["epic_number"],
                task["start_date"],
                task["end_date"],
                task["color"],
                task.get("priority", "Medium"),
                task.get("status", "Not Started"),
                "Yes" if task.get("is_milestone", False) else "No"
            ]
            
            task_start_date = datetime.strptime(task["start_date"], "%Y-%m-%d").date()
            task_end_date = datetime.strptime(task["end_date"], "%Y-%m-%d").date()

            for i, week_start in enumerate(weeks):
                week_end = week_start + timedelta(days=6)
                if max(task_start_date, week_start) <= min(task_end_date, week_end):
                    row_data.append("X") # Marker for task presence
                    
                    col_idx = len(headers) + i + 1
                    cell = ws.cell(row=r_idx + 2, column=col_idx) 
                    
                    fill_color = task["color"].lstrip('#')
                    if len(fill_color) == 6:
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    elif len(fill_color) == 8:
                         cell.fill = PatternFill(start_color=fill_color[2:], end_color=fill_color[2:], fill_type="solid")
                    else:
                        print(f"Warning: Invalid color format for Excel: {task['color']}")
                else:
                    row_data.append("")

            ws.append(row_data)

        for col_idx, col in enumerate(ws.columns):
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width


        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Export Gantt Chart to Excel",
            initialdir=self.file_dialog_initial_dir
        )
        if file_path:
            try:
                wb.save(file_path)
                messagebox.showinfo("Export Successful", f"Gantt chart exported to {file_path}")
                self.update_status(f"Gantt chart exported to {os.path.basename(file_path)}.")
                self.file_dialog_initial_dir = os.path.dirname(file_path)
            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to export to Excel: {e}")

    def export_to_png(self):
        messagebox.showinfo("Export to PNG", "Exporting Canvas to PNG is complex in Tkinter and requires additional libraries (e.g., Pillow) and rendering logic. This feature is not directly implemented in this version.")


if __name__ == "__main__":
    root = tk.Tk()
    app = GanttChartApp(root)
    root.mainloop()
